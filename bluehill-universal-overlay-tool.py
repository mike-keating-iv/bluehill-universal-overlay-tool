# Import dependencies
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import ScatterChart, Reference, Series
import tkinter as tk
import tkinter.scrolledtext as st
from tkinter import filedialog
import os
import pandas as pd
import glob


# Set script directory variable
SCRIPT_DIRECTORY = os.path.dirname(__file__)

# Create openpyxl workbook
wb = openpyxl.Workbook()

# Create worksheet where the chart will be generated
ws = wb.active
ws.title = "Overlay"

# Function to select folder containing .csv files
def select_directory():
    global folder_path
    folder_path = filedialog.askdirectory(initialdir=SCRIPT_DIRECTORY)
    if folder_path:
        print(f"Selected folder: {folder_path}")
        tk.messagebox.showinfo("Folder Selected", f"Folder selected:\n{folder_path}")


# Function to process and chart data from each .csv file
def process_files_in_folder():
    csv_files = [file for file in glob.glob(os.path.join(folder_path,'*.csv'))]
    if csv_files:
        print("Found CSV files in folder:")
        for csv_file in csv_files:
            print(f"-{csv_file}")
            text_area.insert(tk.INSERT, f"{csv_file}\n")
            read_csv_to_excel(csv_file)

        


        # Ask user where to save excel export
        #export_path = filedialog.asksaveasfilename(defaultextension='.xlsx')
        #wb.save(export_path)
        #tk.messagebox.showinfo("Information", f"Excel Overlay Saved. Program will now close.\nSave location:\n{export_path}")
        #root.quit()  
    else:
        tk.messagebox.showinfo("Error", f"Error: No CSV File Found in {folder_path} ")
        print("Error: No CSV File Found")

def export_to_excel():

    # Add logic to chart scatter plots
    # Set active sheet to Overlay
    ws = wb['Overlay']

    # Create Scatter Chart
    chart = ScatterChart(scatterStyle='smoothMarker')
    chart.title = "Tenacity vs. Displacement"
    chart.x_axis.title = 'Elongation (%)'
    chart.y_axis.title = 'Tenacity (gf/den)'
    chart.x_axis.scaling.min = 0
    chart.y_axis.scaling.min = 0
    chart.height = 15
    chart.width = 30

    # Iterate through each sheet and append the data
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Exclude empty first sheet in the series generation
        if ws.title == "Overlay":
            continue
        # Assign x and y data for chart
        x_data = Reference(ws, min_col=4, min_row=2, max_row=ws.max_row)
        y_data = Reference(ws, min_col=5, min_row = 2, max_row=ws.max_row)
        series = Series(y_data,x_data, title_from_data=False, title=sheet)
        series.smooth = True
        chart.series.append(series)
        
    #Set Overlay to Active Sheet and add chart    
    ws = wb['Overlay']
    ws.add_chart(chart, "A1")

    # Ask user where to save excel export
    export_path = filedialog.asksaveasfilename(defaultextension='.xlsx')
    wb.save(export_path)
    tk.messagebox.showinfo("Information", f"Excel Overlay Saved. Program will now close.\nSave location:\n{export_path}")
    root.quit()

def read_csv_to_excel(csv_file):
    global wb
    global guage_length
    global denier

    # Read csv and treat first two rows as header
    df = pd.read_csv(csv_file, header=[0,1])

    # Format header to have no spaces
    df.columns = df.columns.map(lambda h: '  '.join(h).replace(' ', '_'))

    # Calculate displacement based on gauge 
    df['displacement(%)'] = (df['Displacement__(in)']/gauge_length.get()) * 100

    # Calculate tenacity (gf/den) based on denier
    # 1 lbf = 453.59237 gf
    df['tenacity(gf/den)'] = (df['Force__(lbf)']*453.59237) / denier.get()

    # Drop all rows after the row with maximum force
    max_force_index = df['Force__(lbf)'].idxmax()
    df = df.loc[:max_force_index]
    

   
    #Convert df to rows
    rows = dataframe_to_rows(df, index=False)

    #Save data to worksheet and title as csv
    #Insert at first position if it is the first sheed added
    ws = wb.create_sheet(os.path.basename(csv_file))
    for row in rows:
        ws.append(row)
    
# Create the main window 
root = tk.Tk()
root.title("Bluehill Universal Export Overlay Tool")

# Create a label for instructions
intro_instructions_label = tk.Label(root, 
                              text="""1. For each stress strain curve you would like to overlay, choose a .csv file from the autogenerated export folder from Bluehill Universal and copy it into a folder.
                              \n2. Use the button below to select the folder that contains the BlueHill Universal Instron export files.""")
intro_instructions_label.pack(pady=5,padx=10)

# Create a button to open the file dialog
select_button = tk.Button(root, text="Select Folder", command=select_directory)
select_button.pack(padx=10, pady=5)

# Create a label for instructions on gauge length and denier entry
entry_instructions_label = tk.Label(root, text="3.Enter gauge length and denier below.")
entry_instructions_label.pack(pady=10,padx=10)

# Create entry widget to enter Guage Length
gauge_length = tk.DoubleVar()
gauge_entry_label = tk.Label(root, text="Enter Gauge Length (inches):")
gauge_entry_label.pack()
gauge_entry = tk.Entry(root, textvariable=gauge_length)
gauge_entry.delete(0,tk.END)
# Insert 2 inches as default gauge length
gauge_entry.insert(0,"2.0")
gauge_entry.pack()

# Create entry widget to enter Fiber Denier
denier = tk.DoubleVar()
denier_entry_label = tk.Label(root, text="Enter Fiber Denier:")
denier_entry_label.pack()
denier_entry = tk.Entry(root, textvariable=denier)
denier_entry.delete(0,tk.END)
# Insert 40 as default denier
denier_entry.insert(0,"40")
denier_entry.pack()


# Create a label for instructions on processing files to excel
process_files_instructions_label = tk.Label(root, text = "4. Select Process Files in Folder to add curves from selected folder to the chart. \nRepeat with other folders (denier) if needed. \nClick Export to Excel File when finished ")
process_files_instructions_label.pack(pady=10)

# Create a frame for the bottom two buttons
bottom_frame = tk.Frame(root)
bottom_frame.pack(side = 'bottom')

# Create a button to run the process_files_in_folder command
process_button = tk.Button(bottom_frame, text="Process Files in Folder", command=process_files_in_folder)
process_button.pack(side='left', pady=20)

# Create a button to save processed files to excel
export_to_excel_button = tk.Button(bottom_frame, text="Export to Excel File", command=export_to_excel, bg='green')
export_to_excel_button.pack(side='right',padx=10)

# Create scrolled text area to display the processed files
processed_files_label = tk.Label(root, text = "Files Currently Processed:")
processed_files_label.pack()
text_area = st.ScrolledText(root, width = 120, height = 20,font=("Times New Roman",10))
text_area.pack()


# Run the GUI
root.mainloop()